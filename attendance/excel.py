"""
ماژول Export به Excel
"""
from io import BytesIO
from django.http import HttpResponse
from django.db.models import Q
from openpyxl import Workbook
from excelstyler import create_header, create_value, shamsi_date
from excelstyler.headers import create_header_freez
from excelstyler.helpers import excel_description
import jdatetime
from .models import Student, Attendance, Class


def export_students_excel(students_query, class_filter='', search_query=''):
    """
    Export لیست دانش‌آموزان به Excel
    
    Args:
        students_query: QuerySet دانش‌آموزان
        class_filter: فیلتر کلاس
        search_query: فیلتر جستجو
    """
    # دریافت همان فیلترهای صفحه لیست
    students = students_query.filter(is_active=True).select_related('parent', 'class_room', 'class_room__teacher').order_by('-created_at')
    
    if class_filter:
        students = students.filter(class_room_id=class_filter)
    
    if search_query:
        students = students.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(student_id__icontains=search_query)
        )
    
    # آمار کلی
    total_students_count = students.count()
    students_with_class = students.exclude(class_room__isnull=True).count()
    students_without_class = total_students_count - students_with_class
    
    # ایجاد Excel
    output = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "لیست دانش‌آموزان"
    worksheet.sheet_view.rightToLeft = True
    
    # عنوان گزارش
    excel_description(worksheet, 'A1', 'گزارش لیست دانش‌آموزان', size=12, color='red', to_row='L1')
    
    # آمار کلی - باید بعد از row 1 باشد تا با merge تداخل نداشته باشد
    header_list = [
        'تعداد کل دانش‌آموزان',
        'تعداد دانش‌آموزان دارای کلاس',
        'تعداد دانش‌آموزان بدون کلاس',
    ]
    create_header(worksheet, header_list, start_col=2, row=2, height=21, width=22)
    
    value_header_list = [
        total_students_count,
        students_with_class,
        students_without_class,
    ]
    # استفاده از start_col=3, row=2 برای جلوگیری از تداخل با merge
    create_value(worksheet, value_header_list, start_col=3, row=2, border_style='thin')
    
    # هدرهای جدول اصلی
    excel_options = [
        'ردیف',
        'شماره دانش‌آموزی',
        'نام',
        'نام خانوادگی',
        'کلاس',
        'پایه',
        'معلم',
        'نام اولیا',
        'نام خانوادگی اولیا',
        'شماره تماس اولیا',
        'وضعیت',
        'تاریخ ثبت',
    ]
    
    # ایجاد هدر با freeze
    create_header_freez(worksheet, excel_options, start_col=1, row=5, header_row=6, height=17, width=22)
    
    # نوشتن داده‌ها
    l = 6
    m = 1
    for student in students:
        class_name = student.class_room.name if student.class_room else '-'
        grade = student.class_room.grade if student.class_room else '-'
        teacher_name = f"{student.class_room.teacher.first_name} {student.class_room.teacher.last_name}" if student.class_room and student.class_room.teacher else '-'
        status = 'فعال' if student.is_active else 'غیرفعال'
        
        # تبدیل تاریخ به شمسی
        if student.created_at:
            create_date_shamsi = str(shamsi_date(student.created_at.date()))
        else:
            create_date_shamsi = '-'
        
        list1 = [
            m,
            student.student_id if student.student_id else '-',
            student.first_name if student.first_name else '-',
            student.last_name if student.last_name else '-',
            class_name,
            grade,
            teacher_name,
            student.parent.first_name if student.parent else '-',
            student.parent.last_name if student.parent else '-',
            student.parent.phone_number if student.parent else '-',
            status,
            create_date_shamsi,
        ]
        
        # در create_value: row=start_col, column=item+row
        create_value(worksheet, list1, start_col=l, row=1, m=m, border_style='thin', height=22)
        m += 1
        l += 1
    
    # ردیف مجموع
    list2 = [
        'مجموع==>',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
    ]
    create_value(worksheet, list2, start_col=l + 3, row=1, color='green')
    
    workbook.save(output)
    output.seek(0)
    
    # ایجاد response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'لیست_دانش_آموزان_{jdatetime.datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'.encode('utf-8')
    response.write(output.getvalue())
    return response


def export_attendance_excel(attendances_query, date_filter=None, status_filter='', absent_only=False):
    """
    Export لیست حضور و غیاب به Excel
    
    Args:
        attendances_query: QuerySet حضور و غیاب
        date_filter: فیلتر تاریخ (date object)
        status_filter: فیلتر وضعیت
        absent_only: فقط غایب‌ها
    """
    # اعمال فیلترها
    attendances = attendances_query.select_related(
        'student', 'student__parent', 'student__class_room', 'student__class_room__teacher'
    )
    
    if date_filter:
        attendances = attendances.filter(date=date_filter)
    
    if status_filter:
        attendances = attendances.filter(status=status_filter)
    
    if absent_only:
        attendances = attendances.filter(status=Attendance.ABSENT)
    
    attendances = attendances.order_by('-date', 'student')
    
    # آمار کلی
    total_count = attendances.count()
    absent_count = attendances.filter(status=Attendance.ABSENT).count()
    excused_count = attendances.filter(status=Attendance.EXCUSED).count()
    late_count = attendances.filter(status=Attendance.LATE).count()
    sms_sent_count = attendances.filter(sms_sent=True).count()
    
    # ایجاد Excel
    output = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "لیست حضور و غیاب"
    worksheet.sheet_view.rightToLeft = True
    
    # عنوان گزارش
    title = 'گزارش لیست حضور و غیاب'
    if date_filter:
        title += f' - تاریخ: {str(shamsi_date(date_filter))}'
    excel_description(worksheet, 'A1', title, size=12, color='red', to_row='J1')
    
    # آمار کلی
    header_list = [
        'تعداد کل',
        'غایب',
        'غایب موجه',
        'تأخیر',
        'پیامک ارسال شده',
    ]
    create_header(worksheet, header_list, start_col=2, row=2, height=21, width=22)
    
    value_header_list = [
        total_count,
        absent_count,
        excused_count,
        late_count,
        sms_sent_count,
    ]
    create_value(worksheet, value_header_list, start_col=3, row=2, border_style='thin')
    
    # هدرهای جدول اصلی
    excel_options = [
        'ردیف',
        'تاریخ',
        'نام دانش‌آموز',
        'نام خانوادگی دانش‌آموز',
        'شماره دانش‌آموزی',
        'کلاس',
        'پایه',
        'معلم',
        'وضعیت',
        'یادداشت',
        'پیامک ارسال شده',
        'نام اولیا',
        'نام خانوادگی اولیا',
        'شماره تماس اولیا',
    ]
    
    # ایجاد هدر با freeze
    create_header_freez(worksheet, excel_options, start_col=1, row=5, header_row=6, height=17, width=22)
    
    # نوشتن داده‌ها
    l = 6
    m = 1
    for attendance in attendances:
        # تبدیل تاریخ به شمسی
        if attendance.date:
            date_shamsi = str(shamsi_date(attendance.date,in_value=True))
        else:
            date_shamsi = '-'
        
        # وضعیت
        status_display = dict(Attendance.STATUS_CHOICES).get(attendance.status, attendance.status)
        
        # کلاس و معلم
        class_name = attendance.student.class_room.name if attendance.student.class_room else '-'
        grade = attendance.student.class_room.grade if attendance.student.class_room else '-'
        teacher_name = f"{attendance.student.class_room.teacher.first_name} {attendance.student.class_room.teacher.last_name}" if attendance.student.class_room and attendance.student.class_room.teacher else '-'
        
        # پیامک
        sms_status = 'بله' if attendance.sms_sent else 'خیر'
        
        list1 = [
            m,
            date_shamsi,
            attendance.student.first_name if attendance.student.first_name else '-',
            attendance.student.last_name if attendance.student.last_name else '-',
            attendance.student.student_id if attendance.student.student_id else '-',
            class_name,
            grade,
            teacher_name,
            status_display,
            attendance.notes if attendance.notes else '-',
            sms_status,
            attendance.student.parent.first_name if attendance.student.parent else '-',
            attendance.student.parent.last_name if attendance.student.parent else '-',
            attendance.student.parent.phone_number if attendance.student.parent else '-',
        ]
        
        create_value(worksheet, list1, start_col=l, row=1, m=m, border_style='thin', height=22)
        m += 1
        l += 1
    
    # ردیف مجموع
    list2 = [
        'مجموع==>',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
    ]
    create_value(worksheet, list2, start_col=l + 3, row=1, color='green')
    
    workbook.save(output)
    output.seek(0)
    
    # ایجاد response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'لیست_حضور_و_غیاب_{jdatetime.datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'.encode('utf-8')
    response.write(output.getvalue())
    return response


def export_student_attendance_excel(student, attendances_query, status_filter=''):
    """
    Export حضور و غیاب یک دانش‌آموز به Excel
    
    Args:
        student: دانش‌آموز
        attendances_query: QuerySet حضور و غیاب
        status_filter: فیلتر وضعیت
    """
    # اعمال فیلتر
    attendances = attendances_query.order_by('-date')
    
    if status_filter:
        attendances = attendances.filter(status=status_filter)
    
    # آمار کلی
    total_count = Attendance.objects.filter(student=student).count()
    absent_count = Attendance.objects.filter(student=student, status=Attendance.ABSENT).count()
    excused_count = Attendance.objects.filter(student=student, status=Attendance.EXCUSED).count()
    late_count = Attendance.objects.filter(student=student, status=Attendance.LATE).count()
    
    # ایجاد Excel
    output = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "حضور و غیاب دانش‌آموز"
    worksheet.sheet_view.rightToLeft = True
    
    # عنوان گزارش
    title = f'گزارش حضور و غیاب: {student.first_name} {student.last_name}'
    excel_description(worksheet, 'A1', title, size=12, color='red', to_row='H1')
    
    # اطلاعات دانش‌آموز
    student_info = [
        'شماره دانش‌آموزی',
        'کلاس',
        'معلم',
        'اولیا',
        'شماره تماس',
    ]
    create_header(worksheet, student_info, start_col=2, row=2, height=21, width=22)
    
    class_name = f"{student.class_room.grade} {student.class_room.name}" if student.class_room else '-'
    teacher_name = f"{student.class_room.teacher.first_name} {student.class_room.teacher.last_name}" if student.class_room and student.class_room.teacher else '-'
    parent_name = f"{student.parent.first_name} {student.parent.last_name}" if student.parent else '-'
    parent_phone = student.parent.phone_number if student.parent else '-'
    
    student_values = [
        student.student_id,
        class_name,
        teacher_name,
        parent_name,
        parent_phone,
    ]
    create_value(worksheet, student_values, start_col=3, row=2, border_style='thin')
    
    # آمار کلی
    stats_header = [
        'کل رکوردها',
        'غایب',
        'غایب موجه',
        'تأخیر',
    ]
    create_header(worksheet, stats_header, start_col=5, row=2, height=21, width=18)
    
    stats_values = [
        total_count,
        absent_count,
        excused_count,
        late_count,
    ]
    create_value(worksheet, stats_values, start_col=6, row=2, border_style='thin')
    
    # هدرهای جدول اصلی
    excel_options = [
        'ردیف',
        'تاریخ',
        'وضعیت',
        'یادداشت',
        'پیامک ارسال شده',
    ]
    
    # ایجاد هدر با freeze
    create_header_freez(worksheet, excel_options, start_col=1, row=8, header_row=9, height=17, width=25)
    
    # نوشتن داده‌ها
    l = 9
    m = 1
    for attendance in attendances:
        # تبدیل تاریخ به شمسی
        if attendance.date:
            date_shamsi = str(shamsi_date(attendance.date, in_value=True))
        else:
            date_shamsi = '-'
        
        # وضعیت
        status_display = dict(Attendance.STATUS_CHOICES).get(attendance.status, attendance.status)
        
        # پیامک
        sms_status = 'بله' if attendance.sms_sent else 'خیر'
        
        list1 = [
            m,
            date_shamsi,
            status_display,
            attendance.notes if attendance.notes else '-',
            sms_status,
        ]
        
        create_value(worksheet, list1, start_col=l, row=1, m=m, border_style='thin', height=22)
        m += 1
        l += 1
    
    workbook.save(output)
    output.seek(0)
    
    # ایجاد response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    student_name = f'{student.first_name}_{student.last_name}'.replace(' ', '_')
    filename = f'حضور_و_غیاب_{student_name}_{jdatetime.datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'.encode('utf-8')
    response.write(output.getvalue())
    return response

