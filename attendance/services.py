"""
سرویس ارسال پیامک
"""
from io import BytesIO

import openpyxl
import requests
import urllib.parse
from django.conf import settings
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from excelstyler import shamsi_date
from .models import Attendance, Student, Class, Parent, Teacher, UserProfile
from django.contrib.auth.models import User

# تنظیمات سرویس پیامک ساهند
SMS_USERNAME = 'hamedan'
SMS_PASSWORD = 'hamedan12'
SMS_SENDER = '30002501'
SMS_API_URL = 'http://webservice.sahandsms.com/newsmswebservice.asmx/SendPostUrl'


def check_mobile_number(mobile):
    """بررسی صحت شماره موبایل"""
    import re
    if mobile and re.match(r'^09\d{9}$', mobile):
        return True
    return False


def format_persian_date(date_obj):
    """فرمت تاریخ شمسی با استفاده از excelstyler"""
    return str(shamsi_date(date_obj))


def send_sms(phone_number, message):
    """
    ارسال پیامک به شماره تلفن با استفاده از سرویس ساهند
    """
    if not check_mobile_number(phone_number):
        print(f"[SMS] شماره موبایل نامعتبر: {phone_number}")
        return False

    try:
        # انکود کردن پیام برای URL
        encoded_message = urllib.parse.quote(message)
        
        # ساخت URL درخواست
        url = (
            f"{SMS_API_URL}?"
            f"username={SMS_USERNAME}&"
            f"password={SMS_PASSWORD}&"
            f"from={SMS_SENDER}&"
            f"to={phone_number}&"
            f"message={encoded_message}"
        )
        
        response = requests.get(url, timeout=15)
        
        print(f"[SMS] ارسال به {phone_number} - Status: {response.status_code}")
        
        if response.status_code == 200:
            return True
        else:
            print(f"[SMS] خطا در ارسال: {response.text}")
            return False
            
    except Exception as e:
        print(f"[SMS] خطا در ارسال پیامک: {str(e)}")
        return False


def send_absence_sms(attendance, school_name=None):
    """
    ارسال پیامک غیبت/تاخیر به اولیا دانش‌آموز
    """
    if attendance.sms_sent:
        return True  # قبلاً ارسال شده
    
    student = attendance.student
    parent = student.parent
    
    # دریافت نام مدرسه از پروفایل کاربر (اگر وجود داشته باشد)
    if not school_name:
        school_name = "مدرسه"
    elif "مدرسه" not in school_name:
        school_name = f"مدرسه {school_name}"
    
    # تعیین نوع وضعیت
    status_text = ""
    if attendance.status == Attendance.ABSENT:
        status_text = "غیبت غیرموجه"
    elif attendance.status == Attendance.EXCUSED:
        status_text = "غیبت موجه"
    elif attendance.status == Attendance.LATE:
        status_text = "تاخیر"
    else:
        return False  # وضعیت نامعتبر
    
    # تاریخ شمسی
    persian_date = format_persian_date(attendance.date)
    
    # متن پیامک
    message = (
        f"اولیای محترم\n"
        f"فرزند گرامی شما {student.first_name} {student.last_name} "
        f"در تاریخ {persian_date} "
        f"{status_text} داشته است.\n"
        f"با احترام {school_name}"
    )
    
    # ارسال پیامک
    success = send_sms(parent.phone_number, message)
    
    if success:
        attendance.sms_sent = True
        attendance.save(update_fields=['sms_sent'])
    
    return success


def test_sms_simple():
    """
    تابع ساده برای تست ارسال پیامک
    """
    mobile = '09165597588'
    message = 'تست ارسال پیامک از سامانه حضور و غیاب'
    
    if check_mobile_number(mobile):
        try:
            encoded_message = urllib.parse.quote(message)
            url = (
                f"{SMS_API_URL}?"
                f"username={SMS_USERNAME}&"
                f"password={SMS_PASSWORD}&"
                f"from={SMS_SENDER}&"
                f"to={mobile}&"
                f"message={encoded_message}"
            )
            response = requests.get(url, timeout=15)
            print(f"SMS sent successfully. Status: {response.status_code}")
            return f"پیامک با موفقیت ارسال شد. Status: {response.status_code}"
        except Exception as e:
            print(f"Error sending SMS: {str(e)}")
            return f"خطا در ارسال پیامک: {str(e)}"
    else:
        return "شماره موبایل معتبر نیست"



@csrf_exempt
def upload_excel_file(request):
    """
    آپلود فایل اکسل و ایجاد دانش‌آموزان
    فرمت اکسل:
    - یک شیت (یا شیت اول)
    - ردیف اول: هدر (ردیف، نام، نام خانوادگی، کلاس) - نادیده گرفته می‌شود
    - ردیف‌های بعدی: 
      - ستون 0 (A) = ردیف (نادیده گرفته می‌شود)
      - ستون 1 (B) = نام
      - ستون 2 (C) = نام خانوادگی
      - ستون 3 (D) = ID کلاس (عدد - باید کلاس از قبل در سیستم وجود داشته باشد)
    """
    if 'file' not in request.FILES:
        return {
            'success': False,
            'message': 'فایل ارسال نشده است.',
            'created_students': 0,
            'created_parents': 0,
            'errors': []
        }
    
    file = request.FILES['file'].read()
    wb_obj = openpyxl.load_workbook(filename=BytesIO(file))
    
    created_students = 0
    created_parents = 0
    errors = []
    
    # استفاده از شیت اول (یا اولین شیت فعال)
    sheet = wb_obj.active if wb_obj.active else wb_obj[wb_obj.sheetnames[0]]
    sheet_name = sheet.title
    
    # پردازش ردیف‌های شیت
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # رد کردن ردیف اول (هدر)
        if row_idx == 1:
            continue
        
        try:
            # خواندن داده‌ها از ستون‌ها
            # ستون 0 = ردیف (نادیده گرفته می‌شود)
            # ستون 1 = نام
            # ستون 2 = نام خانوادگی
            # ستون 3 = ID کلاس
            first_name = ""
            last_name = ""
            class_id_str = ""
            
            if len(row) > 1 and row[1]:
                first_name = str(row[1]).strip()
            if len(row) > 2 and row[2]:
                last_name = str(row[2]).strip()
            if len(row) > 3 and row[3]:
                class_id_str = str(row[3]).strip()
            
            # اگر ردیف کاملاً خالی است، رد کن
            if not first_name and not last_name and not class_id_str:
                continue
            
            # بررسی اینکه نام و نام خانوادگی وجود داشته باشد
            if not first_name or not last_name:
                errors.append(f"ردیف {row_idx}: نام یا نام خانوادگی خالی است (نام: '{first_name}', نام خانوادگی: '{last_name}')")
                continue
            
            # بررسی اینکه ID کلاس وجود داشته باشد
            if not class_id_str:
                errors.append(f"ردیف {row_idx}: ID کلاس خالی است")
                continue
            
            # تبدیل ID کلاس به عدد
            try:
                class_id = int(class_id_str)
            except ValueError:
                errors.append(f"ردیف {row_idx}: ID کلاس نامعتبر است: '{class_id_str}' (باید عدد باشد)")
                continue
            
            # پیدا کردن کلاس با استفاده از ID
            try:
                class_obj = Class.objects.get(id=class_id, is_active=True)
            except Class.DoesNotExist:
                errors.append(f"ردیف {row_idx}: کلاس با ID '{class_id}' یافت نشد")
                continue
            except Exception as e:
                errors.append(f"ردیف {row_idx}: خطا در پیدا کردن کلاس با ID '{class_id}': {str(e)}")
                continue
            
            # ایجاد یا دریافت اولیا (با همان نام و نام خانوادگی دانش‌آموز)
            parent, parent_created = Parent.objects.get_or_create(
                first_name=first_name,
                last_name=last_name,
                defaults={}
            )
            
            if parent_created:
                created_parents += 1
                print(f"[Excel] اولیا ایجاد شد: {parent}")
            
            # ایجاد دانش‌آموز - استفاده از نام و نام خانوادگی و کلاس برای شناسایی
            # اگر دانش‌آموز با همین نام و نام خانوادگی در همین کلاس وجود داشت، از همان استفاده می‌کنیم
            student = Student.objects.filter(
                first_name=first_name,
                last_name=last_name,
                class_room=class_obj
            ).first()
            
            if student:
                # اگر دانش‌آموز از قبل وجود داشت، فقط کلاس را به‌روزرسانی کن (اگر تغییر کرده)
                if student.class_room != class_obj:
                    student.class_room = class_obj
                    student.save(update_fields=['class_room'])
                    print(f"[Excel] کلاس دانش‌آموز به‌روزرسانی شد: {student}")
            else:
                # ایجاد دانش‌آموز جدید
                student = Student.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    parent=parent,
                    class_room=class_obj,
                    is_active=True
                )
                created_students += 1
                print(f"[Excel] دانش‌آموز ایجاد شد: {student}")
            
            
        except Exception as e:
            errors.append(f"خطا در ردیف {row_idx}: {str(e)}")
            continue
    
    return {
        'success': True,
        'message': f'فایل با موفقیت پردازش شد.',
        'created_students': created_students,
        'created_parents': created_parents,
        'errors': errors
    }


@csrf_exempt
def upload_teachers_excel(request):
    """
    آپلود فایل اکسل و ایجاد معلمان و کاربران
    فرمت اکسل:
    - یک شیت (یا شیت اول)
    - ردیف اول: هدر (ردیف، کدپرسنلی، نام و نام خانوادگی، سمت، کلاس، شماره تماس) - نادیده گرفته می‌شود
    - ردیف‌های بعدی: 
      - ستون 0 (A) = ردیف (نادیده گرفته می‌شود)
      - ستون 1 (B) = کد پرسنلی (نام کاربری)
      - ستون 2 (C) = نام و نام خانوادگی
      - ستون 3 (D) = سمت (نادیده گرفته می‌شود)
      - ستون 4 (E) = کلاس (نادیده گرفته می‌شود)
      - ستون 5 (F) = شماره تماس (رمز عبور - صفر به اولش اضافه می‌شود اگر نداشت)
    """
    if 'file' not in request.FILES:
        return {
            'success': False,
            'message': 'فایل ارسال نشده است.',
            'created_teachers': 0,
            'created_users': 0,
            'errors': []
        }
    
    file = request.FILES['file'].read()
    wb_obj = openpyxl.load_workbook(filename=BytesIO(file))
    
    created_teachers = 0
    created_users = 0
    errors = []
    
    # استفاده از شیت اول (یا اولین شیت فعال)
    sheet = wb_obj.active if wb_obj.active else wb_obj[wb_obj.sheetnames[0]]
    sheet_name = sheet.title
    
    # پردازش ردیف‌های شیت
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # رد کردن ردیف اول (هدر)
        if row_idx == 1:
            continue
        
        try:
            # خواندن داده‌ها از ستون‌ها
            # ستون 0 = ردیف (نادیده گرفته می‌شود)
            # ستون 1 = کد پرسنلی (نام کاربری)
            # ستون 2 = نام و نام خانوادگی
            # ستون 3 = سمت (نادیده گرفته می‌شود)
            # ستون 4 = ID کلاس
            # ستون 5 = شماره تماس
            personnel_code = ""
            full_name = ""
            class_id_str = ""
            phone_number = ""
            
            if len(row) > 1 and row[1]:
                personnel_code = str(row[1]).strip()
            if len(row) > 2 and row[2]:
                full_name = str(row[2]).strip()
            if len(row) > 5 and row[5]:
                phone_number = str(row[5]).strip()
            
            # اگر ردیف کاملاً خالی است، رد کن
            if not personnel_code and not full_name:
                continue
            
            # بررسی اینکه کد پرسنلی وجود داشته باشد
            if not personnel_code:
                errors.append(f"ردیف {row_idx}: کد پرسنلی خالی است")
                continue
            
            # بررسی اینکه نام و نام خانوادگی وجود داشته باشد
            if not full_name:
                errors.append(f"ردیف {row_idx}: نام و نام خانوادگی خالی است")
                continue
            
            # تقسیم نام و نام خانوادگی
            name_parts = full_name.split()
            if len(name_parts) < 2:
                errors.append(f"ردیف {row_idx}: نام و نام خانوادگی باید جدا باشد (نام: '{full_name}')")
                continue
            
            first_name = name_parts[0]
            last_name = ' '.join(name_parts[1:])
            
            # پردازش شماره تلفن - اضافه کردن صفر به اولش اگر نداشت
            if phone_number:
                phone_number = phone_number.strip()
                # اگر شماره تلفن با 0 شروع نمی‌شود، 0 اضافه کن
                if not phone_number.startswith('0'):
                    phone_number = '0' + phone_number
                # بررسی فرمت شماره تلفن
                if not phone_number.startswith('09') or len(phone_number) != 11:
                    errors.append(f"ردیف {row_idx}: شماره تلفن نامعتبر است: '{phone_number}'")
                    continue
            else:
                # اگر شماره تلفن خالی بود، از کد پرسنلی به عنوان شماره تلفن استفاده کن
                phone_number = personnel_code
                if not phone_number.startswith('0'):
                    phone_number = '0' + phone_number
                if len(phone_number) != 11:
                    # اگر طول مناسب نیست، یک شماره پیش‌فرض استفاده کن
                    phone_number = None
            
            # بررسی اینکه کاربر با این کد پرسنلی از قبل وجود نداشته باشد
            if User.objects.filter(username=personnel_code).exists():
                user = User.objects.get(username=personnel_code)
                # به‌روزرسانی اطلاعات کاربر
                user.first_name = first_name
                user.last_name = last_name
                user.save()
                
                # به‌روزرسانی پروفایل
                if hasattr(user, 'profile'):
                    profile = user.profile
                    profile.role = 'teacher'
                    if phone_number:
                        profile.phone_number = phone_number
                    # اگر معلم مرتبط ندارد، ایجاد کن
                    if not profile.teacher:
                        teacher, _ = Teacher.objects.get_or_create(
                            first_name=first_name,
                            last_name=last_name,
                            defaults={'phone_number': phone_number}
                        )
                        profile.teacher = teacher
                    else:
                        teacher = profile.teacher
                        teacher.first_name = first_name
                        teacher.last_name = last_name
                        if phone_number:
                            teacher.phone_number = phone_number
                        teacher.save()
                    profile.save()
                else:
                    # ایجاد پروفایل
                    teacher, _ = Teacher.objects.get_or_create(
                        first_name=first_name,
                        last_name=last_name,
                        defaults={'phone_number': phone_number}
                    )
                    profile = UserProfile.objects.create(
                        user=user,
                        role='teacher',
                        phone_number=phone_number if phone_number else None,
                        teacher=teacher
                    )
                
                print(f"[Excel] کاربر به‌روزرسانی شد: {user.username}")
                continue
            
            # ایجاد کاربر جدید
            # رمز عبور = شماره تلفن (یا کد پرسنلی اگر شماره تلفن نبود)
            password = phone_number if phone_number else personnel_code
            
            try:
                user = User.objects.create_user(
                    username=personnel_code,
                    password=password,
                    first_name=first_name,
                    last_name=last_name
                )
                created_users += 1
                print(f"[Excel] کاربر ایجاد شد: {user.username}")
            except Exception as e:
                errors.append(f"ردیف {row_idx}: خطا در ایجاد کاربر: {str(e)}")
                continue
            
            # ایجاد معلم
            teacher, teacher_created = Teacher.objects.get_or_create(
                first_name=first_name,
                last_name=last_name,
                defaults={'phone_number': phone_number if phone_number else None}
            )
            
            if teacher_created:
                created_teachers += 1
                print(f"[Excel] معلم ایجاد شد: {teacher}")
            
            # ایجاد یا به‌روزرسانی پروفایل
            profile, profile_created = UserProfile.objects.get_or_create(
                user=user,
                defaults={
                    'role': 'teacher',
                    'phone_number': phone_number if phone_number else None,
                    'teacher': teacher
                }
            )
            
            if not profile_created:
                # اگر پروفایل از قبل وجود داشت، به‌روزرسانی کن
                profile.role = 'teacher'
                profile.teacher = teacher
                if phone_number:
                    profile.phone_number = phone_number
                profile.save()
            
        except Exception as e:
            errors.append(f"خطا در ردیف {row_idx}: {str(e)}")
            continue
    
    return {
        'success': True,
        'message': f'فایل با موفقیت پردازش شد.',
        'created_teachers': created_teachers,
        'created_users': created_users,
        'errors': errors
    }
    